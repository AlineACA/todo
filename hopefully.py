from openpyxl import load_workbook

def match_data_between_workbooks(file1, sheet1_name, match_column1, result_column,
                                 file2, sheet2_name, match_column2):
    """
    Compare data between two Excel workbooks and add a column in the first workbook
    to indicate whether the data matches with the second workbook.

    :param file1: Path to the first Excel file.
    :param sheet1_name: Name of the sheet in the first workbook.
    :param match_column1: The column number in the first sheet to match data.
    :param result_column: The column number to insert match results in the first sheet.
    :param file2: Path to the second Excel file.
    :param sheet2_name: Name of the sheet in the second workbook.
    :param match_column2: The column number in the second sheet to match data.
    """

    # Load the workbooks and sheets
    workbook1 = load_workbook(file1)
    sheet1 = workbook1[sheet1_name]

    workbook2 = load_workbook(file2)
    sheet2 = workbook2[sheet2_name]

    # Collect data from the match column in the second sheet
    sheet2_data = {row[match_column2 - 1].value for row in sheet2.iter_rows(min_row=2)}

    # Iterate over the rows in the first sheet and check for matches
    for row in sheet1.iter_rows(min_row=2):
        cell_value = row[match_column1 - 1].value
        if cell_value in sheet2_data:
            row[result_column - 1].value = "Match"
        else:
            row[result_column - 1].value = "No Match"

    # Save the first workbook with the results
    workbook1.save(file1)
    print(f"Results have been saved in '{file1}'.")

# Example usage
match_data_between_workbooks(
    file1='cnpj-N-MEI-2019.xlsx',
    sheet1_name='cnpj2019',
    match_column1=1,  # Column A in the first workbook
    result_column=17, # Column Q in the first workbook where the results will be written
    file2='CNAES Impeditivos.xlsx',
    sheet2_name='CNAES',
    match_column2=11  # Column K in the second workbook
)

